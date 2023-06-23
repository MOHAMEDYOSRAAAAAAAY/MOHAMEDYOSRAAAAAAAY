from pathlib import Path
import datetime
import re
import os
import openpyxl as xl
import subprocess
import win32com.client as  win32  #pip install pywin32
import psutil
# Create output folder
output_dir = Path.cwd() / "Output"
output_dir.mkdir(parents=True, exist_ok=True)

# Connect to outlook
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")

# Connect to folder
#inbox = outlook.Folders("youremail@provider.com").Folders("Inbox")
inbox = outlook.GetDefaultFolder(6)
# https://docs.microsoft.com/en-us/office/vba/api/outlook.oldefaultfolders
# DeletedItems=3, Outbox=4, SentMail=5, Inbox=6, Drafts=16, FolderJunk=23

# Get messages
messages = inbox.Items

for message in messages:

    Lec1 = [1,'Lecture 1', 'lecture1', 'Lecture_1', 'lec1', 'Lec1', 'lecture: 1', 'lecture:1', 'Lec_1', 'c programming basics']
    Lec2 = [2,"Lecture 2" ,"lecture2", "Lecture_2", "lec2" ,"Lec2" ,"lecture: 2" ,"lecture:2", "Lec_2"]
    Lec3 = [3,"Lecture 3" ,"lecture3", "Lecture_3", "lec3" ,"Lec3" ,"lecture: 3" ,"lecture:3", "Lec_3"]
    Lec4 = [4,"Lecture 4" ,"lecture4", "Lecture_4", "lec4" ,"Lec4" ,"lecture: 4" ,"lecture:4", "Lec_4"]
    Lec5 = [5,"Lecture 5" ,"lecture5", "Lecture_5", "lec5" ,"Lec5" ,"lecture: 5" ,"lecture:5", "Lec_5"]
    Lec6 =  [6,"Lecture 6" ,"lecture6", "Lecture_6", "lec6" ,"Lec6" ,"lecture: 6" ,"lecture:6", "Lec_6"]
    
   
    subject = message.Subject
    attachments = message.Attachments
    if   "11104062300" in subject :  #G1
            GropNum = "11104062300"
            Gropnum = 1
            print("Found match in Grop :11104062300")
    elif   "11204062320"  in subject :  #G2
            Gropnum = 2
            GropNum = "11204062320"
            print("Found match in Grop :11204062320")
    elif   "11011062300"  in subject :  #G3
            Gropnum = 3
            GropNum = "11011062300"
            print("Found match in Grop :11011062300")
            
    elif   "11011062301"  in subject :  #G4
                Gropnum = 4
                GropNum = "Found match in Grop :11011062301"
                print("Found match in Grop :11011062301")
    elif   "11011062310" in subject :  #G6
                Gropnum = 6
                GropNum = "11011062310"
                print("Found match in Grop :11011062310")
    elif   "11111062300" in subject :  #G7
                Gropnum = 7
                GropNum = "11111062300"
                print("Found match in Grop :11111062300")   
    elif   "11111062301" in subject :  #G8
                Gropnum = 8
                GropNum = "11111062301"
                print("Found match in Grop :11111062301") 
    elif   "11111062310" in subject :  #G9
                Gropnum = 9
                GropNum = "11111062310"
                print("Found match in Grop :11111062310") 
                
    elif   "11111062311" in subject :  #G10
                Gropnum = 10
                GropNum = "11111062311"
                print("Found match in Grop :11111062311") 
    elif   "11211062300" in subject :  #G11
                Gropnum = 11
                GropNum = "11211062300"
                print("Found match in Grop :11211062300") 
    elif   "11211062301" in subject :  #G12
                Gropnum = 12
                GropNum = "11211062301"
                print("Found match in Grop : 11211062301") 
                
    else : 
            print("Can't Found A matching group")
            GropNum = '0'
            continue
    stdNum = 0
    for i in range (9) :
        if GropNum+"0"+str(i) in subject :
            stdNum = i
            print("The Student Number is " +str(i))    
    if stdNum == 0 :
        for i in range (9,35) :
            if GropNum+str(i) in subject :
                stdNum = i
                print("The Student Number is " +str(i))    
            else :
                print("invaled Student Number ")
    LecNum = '0'
    if GropNum != '0'  and stdNum !=0:           
        listeName = [Lec1,Lec2,Lec3,Lec4,Lec5,Lec6]
        for name in listeName:
            for i in name :
                if str(i) > '9' :
                    if len(name) <= 1:
                        print ("the file of Lec name is empty")    
                    if i in subject:
                        
                        print(i)
                        LecNum  = name[0]
                        break
                        break
                        break
                    elif "Lec1" in str(attachments):
                            LecNum = 1 
                            break
                            break
                            break
                    elif "Lec2" in str(attachments):
                            LecNum = 2 
                            break
                            break
                            break
                    elif "Lec3" in str(attachments):
 
                            LecNum = 3 
                            break
                            break
                            break

                    elif "Lec4" in str(attachments):
                            LecNum = 4 
                            break
                            break
                            break

                    elif "Lec5" in str(attachments):
                            LecNum = 5 
                            break
                            break
                            break

                    elif "Lec6" in str(attachments):
                            LecNum = 6 
                            break
                            break
                            break
       

                  
        if LecNum == '0' :      
                    
                    print  ("Can't fiend a matching Lec Number for "+str(subject )) 
                    print("Please Enter The Lec Number as Lec_1 for example")     
                    LecNumber = input()
                    print("Can You Help us and enter the name that the user used // if No Enter 0 ")
                    LecUser =str(input())
                    if LecUser != '0':
                        if LecNumber in Lec1 :
                               
                               Lec1.append(LecUser) 
                        elif LecNumber in Lec2 :
                               Lec2.append(LecUser) 
                        elif LecNumber in Lec3 :
                               Lec3.append(LecUser) 
                        elif LecNumber in Lec4 :
                               Lec4.append(LecUser) 
                        elif LecNumber in Lec5 :
                               Lec5.append(LecUser) 
                        elif LecNumber in Lec6 :
                               Lec6.append(LecUser) 
                        file = open ("LecName.txt" , 'w+' ,encoding="utf-8")
                        file.write(str(Lec1))
                        file.write('\n')
                        file.write(str(Lec2))
                        file.write('\n')
                        file.write(str(Lec3))
                        file.write('\n')
                        file.write(str(Lec4))
                        file.write('\n')
                        file.write(str(Lec5))
                        file.write('\n')
                        file.write(str(Lec6))
                        file.write('\n')       
                        file.close()       
                                               
      
            
        filename = "G"+str(Gropnum)+".xlsx"
        wb1 = xl.load_workbook(filename)
        ws1 = wb1.worksheets[0]

        mr = ws1.max_row
        mc = ws1.max_column

        filename1 = "G"+str(Gropnum)+"c.xlsx"
        wb2 = xl.load_workbook(filename1)
        ws2 = wb2.active


        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = ws1.cell(row=i, column=j)

                # writing the read value to destination excel file
                ws2.cell(row=i, column=j).value = c.value

        # saving the destination excel file

        sheet = wb2.active

       
        lecnum = str("Lec"+str(LecNum))
        print(lecnum)
        stdnum = str("Student"+str(stdNum))
        # Create separate folder for each message, exclude special characters and timestampe
        current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        #target_folder = output_dir/ GropNum /re.sub('[^0-9a-zA-Z\.]+', '', attachment.FileName)
        target_folder = output_dir / re.sub('[^0-9a-zA-Z]+', '', subject) / stdnum / lecnum
        
        #   filename = re.sub('[^0-9a-zA-Z\.]+', '', attachment.FileName)
        target_folder.mkdir(parents=True, exist_ok=True)
        
        # Save attachments and exclude special
        for attachment in attachments:
            filename = re.sub('[^0-9a-zA-Z\.]+', '', attachment.FileName)
            attachment.SaveAsFile(target_folder / filename)
            print ("\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\"+str(attachment))
            TaskNum = 0
            if "Ass1" in str(attachment):
                            TaskNum = 1 
                            break
                            break
                            break
            elif "Ass2" in str(attachment):
                            TaskNum = 2 
                            break
                            break
                            break
            elif "Ass3" in str(attachment):
 
                            TaskNum = 3 
                            break
                            break
                            break
                            
            if TaskNum ==  0 :
                TaskNum = input ("\nEnter The Assingment Number ")
            
            com = "gcc"+ str(target_folder / filename)+" -o a.exe  "
            output = subprocess.getoutput(com)
            cmd = "gcc "+str(target_folder / filename)+" -o a.exe "

            returned_value = "A"
            print("Displaying the Code")
            log = open(str(target_folder / filename), "r")
            for line in log:
                print(line)
            print("\n\n\n")  
            print("********* Press Enter to Continue *********")
            input()
            
            while returned_value == "A" :

                returned_value = os.system(cmd)  # returns the exit code in unix
   
            if returned_value == 1:
                print("///////////////////////////\n\n\n")
                
                respons =str("Hi There,Thanks for sending Assingment "+str(TaskNum)+" from Lec "+str(LecNum)+"\nBut There is Error in Your code\n"+str(com)+"\nPlease check it and resend it again")
                print(respons)
                valu = "Error"


            else:
                cmd = "a.exe"
                returned_value = "A"
                while returned_value == "A":
                    returned_value = os.system(cmd)
                    print("\n\n\n")
                    valu = int(input("Please Enter the Value: "))

                    print("\n\n\n")
                    respons = str("Hi There,Thanks for sending Assingment " + str(TaskNum) + " from Lec " + str(LecNum) +"\nPlease Keep it up")
                    print(respons)
                  
                    print("\n"
                          "")
            returned_value = "A"

            y=(2+((stdNum-1)*4)+int(TaskNum))
            x=(2+int (LecNum))
            sheet.cell(row=x, column=y).value = 10
            wb2.save(str(filename1))
        #/////////////
            sheet.cell(row=x, column=y).value = valu
            wb2.save(str(filename1))

            filename = "G"+str(Gropnum)+"c.xlsx"
            wb1 = xl.load_workbook(filename)
            ws1 = wb1.worksheets[0]

            mr = ws1.max_row
            mc = ws1.max_column

            filename1 = "G"+str(Gropnum)+".xlsx"



            wb2 = xl.load_workbook(filename1)
            ws2 = wb2.active


            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    # reading cell value from source excel file
                    c = ws1.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    ws2.cell(row=i, column=j).value = c.value

            # saving the destination excel file
            wb2.save(str(filename1))
            
            
               # Open Outlook.exe. Path may vary according to system config
            # Please check the path to .exe file and update below
             
            def send_notification():
                outlook = win32.Dispatch('outlook.application')
                mail = outlook.CreateItem(0)
                mail.To = 'mohamedyosraaaaaaay@gmail.com'
                #message.SenderEmailAddress  
                mail.Subject = str("IMT_Tasks"+subject)
                mail.body = str(respons)
                mail.Send()

            def open_outlook():
                try:
                    subprocess.call(['C:\Program Files (x86)\Microsoft Office\Office14\Outlook.exe'])
                    os.system("C:\Program Files (x86)\Microsoft Office\Office14\Outlook.exe");
                except:
                    print("Outlook didn't open successfully")

            def open_outlook():
                try:
                    subprocess.call(['C:\Program Files (x86)\Microsoft Office\Office14\Outlook.exe'])
                    os.system("C:\Program Files (x86)\Microsoft Office\Office14\Outlook.exe");
                except:
                    print("Outlook didn't open successfully")

            # Checking if outlook is already opened. If not, open Outlook.exe and send email
            for item in psutil.pids():
                p = psutil.Process(item)
                if p.name() == "OUTLOOK.EXE":
                    flag = 1
                    break
                else:
                    flag = 0

            if (flag == 1):
                send_notification()
            else:
                open_outlook()
                send_notification()
                
            print ("Sucsess Proces ")
            #message.Delete()
            
            print("Press Enter to start agine or q to quit")
            Key =     str(input())
            if Key == "q" : break
